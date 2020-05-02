# /usr/bin/env python
# -*- coding:utf-8 -*-
# author: Handsome Lu  time:2020/4/30

import shutil
import os
import cv2
from glob import glob
from PIL import Image
from openpyxl import Workbook
from openpyxl import utils
from openpyxl.styles import  PatternFill
from PyQt5.QtWidgets import QApplication, QMainWindow
import back
import threading
import sys

def color(value):
    temp = []
    temp.append(value[2])
    temp.append(value[1])
    temp.append(value[0])
    digit = list(map(str, range(10))) + list("ABCDEF")
    string = ''
    for i in temp:
      a1 = i // 16
      a2 = i % 16
      string += digit[a1] + digit[a2]
    return string

def del_temp():
    ls = ['temp', 'temp_']
    for x in ls:
        if os.path.exists(x):
            shutil.rmtree(x)
        else:
            print(x + '文件夹已消失，无需删除。')

def save_jpg(path,timeF,n):   #存储路径，提取图片的间隔帧数，缩小倍数
    cap = cv2.VideoCapture(path)
    name = 0
    i=0
    if os.path.exists('temp_'):
        pass
    else:
        os.makedirs('temp_')
    while(cap.isOpened()):
        i=i+1
        ret, frame = cap.read()
        if(i%timeF == 0):
            if ret==True:
                name += 1
                cv2.imwrite('temp_/%03d.jpg'% name,frame)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
            else:
                break
    cap.release()
    cv2.destroyAllWindows()

    img_path = glob("temp_/*.jpg")
    path_save = "temp/"
    if os.path.exists('temp'):
        pass
    else:
        os.makedirs('temp')
    a = range(0, len(img_path))
    i = 0
    for file in img_path:
        name = os.path.join(path_save, "%03d.jpg" % a[i])
        im = Image.open(file)
        z = im.size[0] / n
        im.thumbnail((z, z))
        im.save(name, 'JPEG')
        i += 1

#所有图存入一个文件   暂时问题 :存储文件略大 最后打开就会错误
def jpg_wb(save_name):
    wb = Workbook()
    img_path = glob("temp/*.jpg")
    name = 0
    for file in img_path:
        exec('sheet' + str(name) + '=wb.create_sheet(str(name))')
        x = 1
        img = cv2.imread(file)
        for i in img:
            exec('sheet' + str(name) + '.row_dimensions[x].height =20')
            h = 1
            for j in i:
                exec('sheet' + str(name) + '.column_dimensions[utils.get_column_letter(h)].width =5')
                fill = PatternFill("solid", fgColor=color(j))
                exec('sheet' + str(name)+'.cell(row=x,column=h).fill=fill')
                h += 1
            x += 1
        name += 1
    # #保存文件
    ws = wb["Sheet"]
    wb.remove(ws)
    wb.save(save_name + '.xlsx')

#一张图一个文件
def jpg_wb_2(save_name):
    img_path = glob("temp/*.jpg")
    if os.path.exists(save_name):
        pass
    else:
        os.makedirs(save_name)
    name = 0
    for file in img_path:
        name += 1
        wb = Workbook()
        sheet = wb.create_sheet('sheet')
        x = 1
        img = cv2.imread(file)
        for i in img:
            sheet.row_dimensions[x].height =20
            h = 1
            for j in i:
                sheet.column_dimensions[utils.get_column_letter(h)].width =5
                fill = PatternFill("solid", fgColor=color(j))
                sheet.cell(row=x,column=h).fill=fill
                h += 1
            x += 1
        # #保存文件
        ws = wb["Sheet"]
        wb.remove(ws)
        wb.save(save_name + '/%03d.xlsx'%name)


def work():
    combobox = ui.comboBox.currentIndex()
    In = ui.lineEdit.text()
    Out = ui.lineEdit_2.text()
    timeF = ui.spinBox_2.value()
    Red = ui.spinBox.value()
    del_temp()
    save_jpg(In, timeF, Red)
    if combobox == 0:
        jpg_wb_2(Out)
    elif combobox == 1:
        jpg_wb(Out)
    del_temp()
    ui.pushButton.setDisabled(0)
    ui.pushButton.setText('开始')
def active():
    t = threading.Thread(target=work)
    ui.pushButton.setDisabled(1)
    ui.pushButton.setText('等待')
    t.start()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = back.Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    ui.pushButton.clicked.connect(active)
    sys.exit(app.exec_())